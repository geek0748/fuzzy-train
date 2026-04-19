## Этот файл - пример моей работы с OSR файлов переписи насления Швейцарских кантонов


import os
import cv2
import numpy as np
import pytesseract
from pdf2image import convert_from_path
import re
import pandas as pd
from openpyxl import load_workbook

from img2table.document import Image as Img2TabImg
def adjust_brightness_contrast_gamma(image, alpha=1, beta=0, gamma=1):
    adjusted = cv2.convertScaleAbs(image, alpha=alpha, beta=beta)
    inv_gamma = 1.0 / gamma
    table = np.array([((i / 255.0) ** inv_gamma) * 255
                      for i in np.arange(256)]).astype("uint8")
    return cv2.LUT(adjusted, table)



PDF_FILE = r"C:\Users\Артур\Downloads\RA-Swissdata\RA-Swissdata\RA-Swissdata\Pdf files\1910_Short.pdf"
OUTPUT_DIR = r"C:\temp"

pytesseract.pytesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"


POPPLER_PATH = r"C:\Users\Артур\Downloads\Release-24.08.0-0\poppler-24.08.0\Library\bin"

pd.set_option("display.max_colwidth", None)
pd.set_option("display.width", None)
pd.set_option("display.max_rows", None)
pd.set_option("display.max_columns", None)


for i in range(66, 184, 2):

#     # 1 страница верх
#     page_0 = pages[i]
#
#     # Конвертируем PIL → OpenCV
#     cv_img = cv2.cvtColor(np.array(page_0), cv2.COLOR_RGB2BGR)
#     h, w = cv_img.shape[:2]
#     x1 = 620  # начало предпоследнего столбца
#     x2 = 900  # конец последнего столбца
#     y1 = 550  # первая строка сразу ПОСЛЕ верхней горизонтальной полосы
#     y2 = 1850  # последняя строка сразу НАД нижней горизонтальной полосой
#     crop = cv_img[y1:y2, x1:x2]
#     print("crop shape:", crop.shape if crop is not None else "None")
#     out_path = os.path.join(OUTPUT_DIR, f"up_page_{i}.png")
#     ok = cv2.imwrite(out_path, crop)
#     print("cv2.imwrite returned:", ok)
#
#     # 2 страница верх
#
#     page_1 = pages[i+1]
#     # Конвертируем PIL → OpenCV
#     cv_img = cv2.cvtColor(np.array(page_1), cv2.COLOR_RGB2BGR)
#     h, w = cv_img.shape[:2]
#     print('original page Y:', h, 'X:', w)
#     x1 = 210 # начало предпоследнего столбца
#     x2 = 370  # конец последнего столбца
#     y1 = 550 # первая строка сразу ПОСЛЕ верхней горизонтальной полосы
#     y2 = 1850  # последняя строка сразу НАД нижней горизонтальной полосой
#     crop = cv_img[y1:y2, x1:x2]
#     print("crop shape:", crop.shape if crop is not None else "None")
#     os.makedirs(OUTPUT_DIR, exist_ok=True)  # создаст папку, если её нет
#     out_path = os.path.join(OUTPUT_DIR, f"up_page_{i+1}.png")
#     ok = cv2.imwrite(out_path, crop)
#     print("cv2.imwrite returned:", ok)
#
# #-----------------------------------------------------------------------
#
#     # 1 страница низ
#     page_0 = pages[i]
#
#     # Конвертируем PIL → OpenCV
#     cv_img = cv2.cvtColor(np.array(page_0), cv2.COLOR_RGB2BGR)
#     h, w = cv_img.shape[:2]
#     x1 = 620  # начало предпоследнего столбца
#     x2 = 915  # конец последнего столбца
#     y1 = 1920  # первая строка сразу ПОСЛЕ верхней горизонтальной полосы
#     y2 = 3200  # последняя строка сразу НАД нижней горизонтальной полосой
#     crop = cv_img[y1:y2, x1:x2]
#     print("crop shape:", crop.shape if crop is not None else "None")
#     out_path = os.path.join(OUTPUT_DIR, f"down_page_{i}.png")
#     ok = cv2.imwrite(out_path, crop)
#     print("cv2.imwrite returned:", ok)
#
#     # 2 страница низ
#
#     page_1 = pages[i+1]
#     # Конвертируем PIL → OpenCV
#     cv_img = cv2.cvtColor(np.array(page_1), cv2.COLOR_RGB2BGR)
#     h, w = cv_img.shape[:2]
#     print('original page Y:', h, 'X:', w)
#     x1 = 210 # начало предпоследнего столбца
#     x2 = 370  # конец последнего столбца
#     y1 = 1880 # первая строка сразу ПОСЛЕ верхней горизонтальной полосы
#     y2 = 3100  # последняя строка сразу НАД нижней горизонтальной полосой
#     crop = cv_img[y1:y2, x1:x2]
#     print("crop shape:", crop.shape if crop is not None else "None")
#     os.makedirs(OUTPUT_DIR, exist_ok=True)  # создаст папку, если её нет
#     out_path = os.path.join(OUTPUT_DIR, f"down_page_{i+1}.png")
#     ok = cv2.imwrite(out_path, crop)
#     print("cv2.imwrite returned:", ok)

    def up_page_to_df(page_num):
        img_path = f"C:/temp/up_page_{page_num}.png"
        img      = cv2.imread(img_path)
        data = pytesseract.image_to_data(
            img,
            lang='deu',
            config='--psm 6',
            output_type=pytesseract.Output.DATAFRAME
        )
        nums = data[
            data['text'].fillna('').str.strip().str.contains(r'^[\d\-–—]+$')
        ].copy()
        def to_int(s):
            s = s.strip()
            return int(s) if re.fullmatch(r'\d+', s) else 0
        # группировка по строкам
        nums['row'] = (nums['top'] // 35).astype(int)
        rows = {}
        for _, r in nums.iterrows():
            col = 'left' if r['left'] < 150 else 'right'
            rows.setdefault(r['row'], {})[col] = to_int(r['text'])
        # превращаем словарь в DataFrame
        df = pd.DataFrame([
            [rows[r].get('left', 0), rows[r].get('right', 0)]
            for r in sorted(rows)
        ], columns=['left_col', 'right_col'])
        df = df[~((df["left_col"] == 0) & (df["right_col"] == 0))]
        return df
    def down_page_to_df(page_num):
        img_path = f"C:/temp/down_page_{page_num}.png"
        img      = cv2.imread(img_path)
        data = pytesseract.image_to_data(
            img,
            lang='deu',
            config='--psm 6',
            output_type=pytesseract.Output.DATAFRAME
        )
        # только цифры/прочерки
        nums = data[
            data['text'].fillna('').str.strip().str.contains(r'^[\d\-–—]+$')
        ].copy()
        def to_int(s):
            s = s.strip()
            return int(s) if re.fullmatch(r'\d+', s) else 0
        # группировка по строкам
        nums['row'] = (nums['top'] // 35).astype(int)
        rows = {}
        for _, r in nums.iterrows():
            col = 'left' if r['left'] < 150 else 'right'
            rows.setdefault(r['row'], {})[col] = to_int(r['text'])
        # превращаем словарь в DataFrame
        df = pd.DataFrame([
            [rows[r].get('left', 0), rows[r].get('right', 0)]
            for r in sorted(rows)
        ], columns=['left_col', 'right_col'])
        df = df[~((df["left_col"] == 0) & (df["right_col"] == 0))]
        return df

    up_df0 = up_page_to_df(i)   # первая таблица
    up_df1 = up_page_to_df(i+1)   # первая таблица

    if up_df0.iloc[0][ 'left_col'] * 10 <= up_df0.iloc[1][ 'left_col']:
        up_df0 = up_df0.iloc[1:].reset_index(drop=True)
    if up_df0.iloc[0]['left_col'] >= up_df0.iloc[1]['left_col'] * 6 or up_df0.iloc[0]['right_col'] >= up_df0.iloc[1][ 'right_col'] * 6:
        up_df0 = up_df0.iloc[1:].reset_index(drop=True)
    if up_df1.iloc[0][ 'left_col'] * 10 <= up_df1.iloc[1][ 'left_col']:
        up_df1 = up_df1.iloc[1:].reset_index(drop=True)
    if up_df1.iloc[0]['left_col'] >= up_df1.iloc[1]['left_col'] * 6 or up_df1.iloc[0][ 'right_col'] >= up_df1.iloc[1][ 'right_col'] * 6:
        up_df1 = up_df1.iloc[1:].reset_index(drop=True)

    up_df0['left_col'] = up_df0[['left_col', 'right_col']].max(axis=1)

    if up_df0.iloc[0]['left_col'] > up_df0.iloc[1]['left_col'] * 6:
        up_df0.iloc[0]['left_col'] = int(up_df0.loc[p - 1, 'left_col'] / 6)
    for p in range(2, len(up_df0)-1):
        if up_df0.iloc[p-1][ 'left_col'] > up_df0.iloc[p-2][ 'left_col'] * 6 and up_df0.iloc[p-1][ 'left_col'] > up_df0.iloc[p][ 'left_col']:
            up_df0.loc[p - 1, 'left_col'] = int(up_df0.loc[p - 1, 'left_col'] / 10)

    if up_df1.iloc[0]['left_col'] > up_df1.iloc[1]['left_col'] * 6:
        up_df1.iloc[0]['left_col'] = int(up_df1.loc[p - 1, 'left_col'] / 6)
    for p in range(2, len(up_df1)-1):
        if up_df1.iloc[p-1][ 'left_col'] > up_df1.iloc[p-2][ 'left_col'] * 6 and up_df1.iloc[p-1][ 'left_col'] > up_df1.iloc[p][ 'left_col']:
            up_df1.loc[p - 1, 'left_col'] = int(up_df1.loc[p - 1, 'left_col'] / 10)
    print(up_df0)
    print(up_df1)
    file_path = r"C:\Users\Артур\Downloads\RA-Swissdata\RA-Swissdata\RA-Swissdata\Data District Population 1860-1930.xlsx"

    # собираем список значений
    values = [
        up_df0['left_col'].sum(),
        up_df0['right_col'].sum(),
        " ",
        '',
        up_df1['left_col'].sum(),
        up_df0['left_col'][0],
        up_df0['left_col'][1],
        up_df0['left_col'][2],
        '',
        up_df0['left_col'].iloc[3:12].sum(),
        up_df0['left_col'].iloc[13:].sum(),
        up_df1['left_col'][0],
        up_df1['left_col'][1],
        up_df1['left_col'][2],
        up_df1['left_col'].iloc[13:].sum(),
        '',
        up_df1['left_col'].iloc[4:7].sum(),
    ]

    wb = load_workbook(file_path)
    ws = wb['1910']

    for col_offset, val in enumerate(values, start=2):  # 2 = столбец B
        ws.cell(row=5 + i, column=col_offset, value=val)

    wb.save(file_path)
    print(
        f"Данные о таблице {i + 1} записаны в строку {5 + i} листа 1900, i = {i}")

    # ------------------------------------------------------------------------------------------------------------------------------------

    down_df0 = down_page_to_df(i)   # первая таблица
    down_df1 = down_page_to_df(i+1)   # первая таблица

    if down_df0.iloc[0][ 'left_col'] * 10 <= down_df0.iloc[1][ 'left_col']:
        down_df0 = down_df0.iloc[1:].reset_index(drop=True)
    if down_df0.iloc[0]['left_col'] >= down_df0.iloc[1]['left_col'] * 6 or down_df0.iloc[0][ 'right_col'] >= down_df0.iloc[1][ 'right_col'] * 6:
        down_df0 = down_df0.iloc[1:].reset_index(drop=True)
    if down_df1.iloc[0][ 'left_col'] * 10 <= down_df1.iloc[1][ 'left_col']:
        down_df1 = down_df1.iloc[1:].reset_index(drop=True)
    if down_df1.iloc[0]['left_col'] >= down_df1.iloc[1]['left_col'] * 6 or down_df1.iloc[0][ 'right_col'] >= down_df1.iloc[1][ 'right_col'] * 6:
        down_df1 = down_df1.iloc[1:].reset_index(drop=True)

    down_df0['left_col'] = down_df0[['left_col', 'right_col']].max(axis=1)

    if down_df0.iloc[0]['left_col'] > down_df0.iloc[1]['left_col'] * 6:
        down_df0.iloc[0]['left_col'] = int(down_df0.loc[p - 1, 'left_col'] / 6)
    for p in range(2, len(down_df0)-1):
        if down_df0.iloc[p-1][ 'left_col'] > down_df0.iloc[p-2][ 'left_col'] * 6 and down_df0.iloc[p-1][ 'left_col'] > down_df0.iloc[p][ 'left_col']:
            down_df0.loc[p - 1, 'left_col'] = int(down_df0.loc[p - 1, 'left_col'] / 10)

    if down_df1.iloc[0]['left_col'] > down_df1.iloc[1]['left_col'] * 6:
        down_df1.iloc[0]['left_col'] = int(down_df1.loc[p - 1, 'left_col'] / 6)
    for p in range(2, len(down_df1)-1):
        if down_df1.iloc[p-1][ 'left_col'] > down_df1.iloc[p-2][ 'left_col'] * 6 and down_df1.iloc[p-1][ 'left_col'] > down_df1.iloc[p][ 'left_col']:
            down_df1.loc[p - 1,'left_col'] = int(down_df1.loc[p-1, 'left_col']/10)

    print(down_df0)
    print(down_df1)
    file_path = r"C:\Users\Артур\Downloads\RA-Swissdata\RA-Swissdata\RA-Swissdata\Data District Population 1860-1930.xlsx"

    # собираем список значений
    values = [
        down_df0['left_col'].sum(),
        down_df0['right_col'].sum(),
        '',
        '',
        down_df1['left_col'].sum(),
        down_df0['left_col'][0],
        down_df0['left_col'][1],
        down_df0['left_col'][2],
        '',
        down_df0['left_col'].iloc[3:12].sum(),
        down_df0['left_col'].iloc[13:].sum(),
        down_df1['left_col'][0],
        down_df1['left_col'][1],
        down_df1['left_col'][2],
        down_df1['left_col'].iloc[13:].sum(),
        '',
        down_df1['left_col'].iloc[4:7].sum(),
    ]

    wb = load_workbook(file_path)
    ws = wb['1910']

    # пишем по горизонтали начиная с B1
    for col_offset, val in enumerate(values, start=2):  # 2 = столбец B
        ws.cell(row=6 + i, column=col_offset, value=val)

    wb.save(file_path)
    print(
        f"Данные о таблице {i + 2} записаны в строку {6 + i}  листа 1910, i = {i}")
